from openpyxl import load_workbook

# Load the Excel file
excel_file_path = "./test.xlsx"
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Iterate through rows starting from row 5 (assuming row 1 is the header)
for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
    id_value = row[0]  # Assuming id is in the first column

    if id_value == "0001":
        # Update the value in the third column (column index is 3)
        cell = sheet.cell(row=row_number, column=3)
        cell.value = "test"

# Save the changes
workbook.save(excel_file_path)
print("Values updated in Excel.")
